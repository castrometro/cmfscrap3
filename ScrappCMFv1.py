from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# Fecha de hoy
fecha_de_hoy = time.strftime("%d/%m/%Y")

# Fecha de ayer ajustada a la hora local
fecha_de_ayer = time.strftime("%d/%m/%Y", time.localtime(time.time() - 86400))

# Calcular el último viernes
dia_actual = datetime.now().weekday()  # Lunes es 0, Domingo es 6
# Si hoy es lunes (0), martes (1), miércoles (2), jueves (3) o viernes (4), necesitamos un cálculo especial
# Para viernes a domingo (4-6), simplemente restamos los días hasta el viernes
desplazamiento_hasta_viernes = (dia_actual - 4) % 7
ultimo_viernes = datetime.now() - timedelta(days=desplazamiento_hasta_viernes)
fecha_ultimo_viernes = ultimo_viernes.strftime("%d/%m/%Y")


def agrupar_entidades(df):
  # Lista para almacenar los diccionarios con los datos de cada entidad
    datos_agrupados = []
   # Agrupar el DataFrame por 'Entidad' y concatenar las 'Materias' y 'Enlaces'
    for entidad, group in df.groupby('Entidad'):
        # Unir las materias y enlaces con un salto de línea HTML
        materias_html = '<br>'.join(group['Materia'])
        enlaces_html = '<br>'.join([f'<a href="{link}">Ver Enlace</a>' for link in group['Enlace'].tolist()])
     
        # Añadir el diccionario de datos de la entidad actual a la lista
        datos_agrupados.append({
        'Entidad': entidad,
        'Materia': materias_html,
        'Enlace': enlaces_html
        })
    # Crear el DataFrame agrupado directamente de la lista de diccionarios
    df_agrupado = pd.DataFrame(datos_agrupados, columns=['Entidad', 'Materia', 'Enlace'])
    # Obtener la ruta del directorio del script actual
    directorio_actual = os.path.dirname(os.path.abspath(__file__))

    # Construir la ruta completa del archivo Excel
    nombre_archivo = 'hechos_esenciales_agrupados.xlsx'
    ruta_completa_archivo = os.path.join(directorio_actual, nombre_archivo)
    # Guardar el DataFrame en la ruta construida
    df_agrupado.to_excel(ruta_completa_archivo, index=False)

    return df_agrupado


def enviar_correo(df_agrupado, remitente, contraseña, destinatario, asunto):
    # Convertir el DataFrame a HTML
    html_df = df_agrupado.to_html(escape=False, index=False)

    # HTML personalizado para el cuerpo del correo
    html_correo = f"""
    <html>
        <head>
            <style>
                body {{
                    font-family: 'Arial', sans-serif;
                    margin: 10px;
                }}
                table {{
                    border-collapse: collapse;
                    width: 100%;
                }}
                th, td {{
                    border: 1px solid #dddddd;
                    text-align: left;
                    padding: 8px;
                }}
                th {{
                    background-color: #aa0404;
                    color: white;
                }}
                .footer {{
                    margin-top: 20px;
                    font-size: 0.8em;
                }}
            </style>
        </head>
        <body>
            <h2>Hechos Esenciales</h2>
            <p>Se adjuntan los hechos esenciales más importantes del día de ayer</p>
            <!-- DataFrame HTML-->
            {html_df}
            <p class="footer">Este es un correo automatizado, por favor no responda directamente.</p>
            <!-- Imagen adjunta -->
            <img src="https://gkpb.com.br/wp-content/uploads/2018/03/novo-logo-santander-fundo-vermelho.jpg" alt="Imagende ejemplo" width="150">
        </body>
    </html>
    """



    # Crear el mensaje
    mensaje = MIMEMultipart()
    mensaje['From'] = remitente
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto

    # Adjuntar el DataFrame en HTML al correo
    mensaje.attach(MIMEText(html_correo, 'html'))
    servidor = None
    
    try:
        print("Conectando al servidor...")
        servidor = smtplib.SMTP('smtp.office365.com', 587)
        print("Conectado al servidor.")
    except Exception as e:
        print(f"Ocurrió un error al conectar al servidor: {e}")
        return
    try:
        print("Iniciando sesión...")
        servidor.starttls()
        servidor.login(remitente, contraseña)
        print("Sesión iniciada.")
    except Exception as e:
        print(f"Ocurrió un error al iniciar sesión: {e}")
        return
    try:
        print("Enviando correo...")
        servidor.sendmail(mensaje['From'], mensaje['To'], mensaje.as_string())
        print("Correo enviado.")
        servidor.quit()
        print("Conexión cerrada.")
        return True

    except Exception as e:
        print(f"Ocurrió un error al enviar el correo: {e}")
        return False

def actualizar_y_agregar_a_df(archivo='hechos_esenciales.xlsx'):
  libro = openpyxl.load_workbook(archivo)
  hoja = libro.active
   
  # DataFrame vacío con las columnas específicas a utilizar
  df = pd.DataFrame(columns=['Fecha', 'Hora', 'ID', 'Entidad', 'Materia', 'Enlace'])
   
  # Lista para almacenar temporalmente las filas a agregar
  filas_para_agregar = []
   
  for indice, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
    if fila[-1] == 'N': # Si el valor en la columna "ENVIADO(Y/N)" es "N"
      # Agregar la fila a la lista temporal
      filas_para_agregar.append(fila[:-1])

  # Si hay filas para agregar, convertirlas en un DataFrame y concatenar con el df principal
  if filas_para_agregar:
    df_temp = pd.DataFrame(filas_para_agregar, columns=df.columns)
    df = pd.concat([df, df_temp], ignore_index=True)
   
  libro.save(archivo)
  libro.close()
   
  # Mantener solo las columnas "Entidad", "Materia", "Enlace" en el DataFrame
  df = df[['Entidad', 'Materia', 'Enlace']]
   
  return df


def crear_excel():
    if not os.path.exists('hechos_esenciales.xlsx'):
        archivo = 'hechos_esenciales.xlsx'
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = 'Hechos Esenciales'
        hoja.append(['Fecha', 'Hora', 'ID', 'Entidad', 'Materia', 'Enlace', 'ENVIADO(Y/N)'])
        libro.save(archivo)
        libro.close()
    else:
        print('El archivo "hechos_esenciales.xlsx" ya existe.')
    

def añadir_a_excel(datos):
    filas_agregadas = 0
    archivo = 'hechos_esenciales.xlsx'
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active
    for fila in datos:
        if fila[0] == fecha_de_ayer or fila[0] == fecha_ultimo_viernes: #Si la fecha es la de ayer o la del último viernes
            if fila[2] not in [celda.value for celda in hoja['C']]: #Si el ID no está en la columna C
                if fila[3].lower().find('banco') == -1: #Si la entidad no es un banco
                    if (fila[3].lower().find('tanner') != -1 or fila[3].lower().find('forum') != -1) and fila[4] == ('Colocación de valores en mercados internacionales y/o nacionales'): #Si la entidad es Tanner o Forum y la materia es colocación de valores
                        print('Agregando fila:', fila)
                        fila.append('N')
                        hoja.append(fila)
                        filas_agregadas += 1
                    else:
                        print('La entidad', fila[3], 'no cumple con los requisitos.')
                        print('-----------------------------')
                else: #Si la entidad es un banco
                        fila.append('N')
                        hoja.append(fila)
                        filas_agregadas += 1

            else:
                print('El ID', fila[2], 'ya está en el archivo.')
                print('Nombre:', fila[3])   
                print('-----------------------------')
        else:
            print('La fecha no es la de ayer o la de hoy.')
            print('ID:', fila[2])
            print('Fecha:', fila[0])
            print('-----------------------------')

    print ('Filas agregadas: ', filas_agregadas)
    libro.save(archivo)
    libro.close()
    return 
 


def accederyobtenerdf():
    print('ACCEDIENDO A CMF.....')
    service = Service()

    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--disable-extensions')

    driver = webdriver.Chrome(service=service, options=options)
    driver.get('https://www.cmfchile.cl/portal/principal/613/w3-channel.html')

    css_selector = 'button.btn.btn-outline-secondary.btn-sm'

    WebDriverWait(driver,20)
    time.sleep(5)

    #----VENTANA INICIAL--------
    #esperar a que el elemento sea clickeable, luego hacer click
   # WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, css_selector))).click()
   # time.sleep(2)

    #----SCROL DOWN----------
    div_row_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.ntg-box-mb.animar.tab-pills-cmf")))
    driver.execute_script("arguments[0].scrollIntoView(true);",div_row_element)
    time.sleep(2)

    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div[3]/div/div[1]/div[1]/div/div/div[1]/div/table/tbody')))
    time.sleep(2)


    #-----EXTRAER DATOS--------
    print('....EXTRAYENDO DATOS.....')
    tabla = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[3]/div/div[1]/div[1]/div/div/div[1]/div/table')
    filas = tabla.find_elements(By.TAG_NAME, "tr")
    datos = []
    for i in range(3, len(filas)):
        elemento = filas[i].find_elements(By.TAG_NAME, "td")
        fecha_y_hora = elemento[0].text
        fecha = fecha_y_hora.split(' ')[0]
        hora = fecha_y_hora.split(' ')[1]
        id = elemento[1].text
        entidad = elemento[2].text
        materia = elemento[3].text
        enlace = elemento[1].find_element(By.TAG_NAME, "a")
        #print('fecha:',fecha)
        #print('hora:',hora)
        #print('id:',id)
        #print('entidad:',entidad)
        #print('materia:',materia)
        #print('enlace:', enlace.get_attribute('href'))
        #print('-----------------------------')
        fila = [fecha, hora, id, entidad, materia, enlace.get_attribute('href')]
        datos.append(fila)

    añadir_a_excel(datos)
    print('....FIN EXTRACCIÓN.....')
    return 0

def marcar_filas_enviadas(df, archivo='hechos_esenciales.xlsx'):
    # Marcar las filas del DataFrame como enviadas
    df['ENVIADO(Y/N)'] = 'Y'
    
    # Cargar el libro y seleccionar la hoja activa
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active
    
    # Actualizar las celdas en el Excel para reflejar las filas enviadas
    for indice in range(2, hoja.max_row + 1):  
        if hoja.cell(row=indice, column=7).value == 'N':  # 'ENVIADO(Y/N)' en la columna 7
            hoja.cell(row=indice, column=7, value='Y')
    
    # Guardar y cerrar el libro
    libro.save(archivo)
    libro.close()

def main():
    crear_excel()
    accederyobtenerdf()
    df = actualizar_y_agregar_a_df()
    df_agrupado = agrupar_entidades(df)
    if df_agrupado.empty:
        print('No hay hechos esenciales para enviar.')
        return
    else:
        print(df)
        print('....EXCEL ACTUALIZADO.....')
        #enviar_correo(df_agrupado,'felipe.salles@santander.cl', 'Boletín de Hechos Esenciales')
        #enviar_correo(df_agrupado,'pablo.castro@servexternos.santander.cl', 'Boletín de Hechos Esenciales')
        #enviar_correo(df_agrupado,'emiliano.muratore@santander.cl', 'Boletín de Hechos Esenciales') 
        #enviar_correo(df_agrupado,'javier.torrealba@santander.cl', 'Boletín de Hechos Esenciales')
        remitente = ' ' #correo del remitente: juanito@outlook.com
        contraseña = ' ' #contraseña del remitente: 123456
        destinatario = ' ' #correo del destinatario: juan@gmail.com
        envio_exitoso = enviar_correo(df_agrupado,remitente, contraseña, destinatario, 'Boletín de Hechos Esenciales')
        if envio_exitoso:
            marcar_filas_enviadas(df)

    
if __name__ == "__main__":
    main()